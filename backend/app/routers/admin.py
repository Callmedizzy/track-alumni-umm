import io
import math
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Request, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.models import AlumniBase, AlumniContact, AlumniCareer, User, AuditLog
from app.schemas import AuditLogOut, PaginatedAuditLog
from app.security import get_current_user, require_admin, log_action

router = APIRouter(tags=["Admin & Export"])


# ─── GET /export/excel ────────────────────────────────────────────────────────

@router.get("/export/excel")
def export_excel(
    request: Request,
    fakultas: Optional[str] = Query(None),
    prodi: Optional[str] = Query(None),
    tahun: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl tidak terinstall.")

    query = db.query(AlumniBase).options(
        joinedload(AlumniBase.contact),
        joinedload(AlumniBase.career),
    )
    if fakultas:
        query = query.filter(AlumniBase.fakultas.ilike(f"%{fakultas}%"))
    if prodi:
        query = query.filter(AlumniBase.prodi.ilike(f"%{prodi}%"))
    if tahun:
        query = query.filter(func.extract("year", AlumniBase.tgl_lulus) == tahun)

    records = query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Alumni"

    headers = [
        "Nama", "NIM", "Tahun Masuk", "Tanggal Lulus", "Fakultas", "Program Studi",
        "Email", "No HP", "LinkedIn", "Instagram", "Facebook", "TikTok",
        "Tempat Kerja", "Alamat Kerja", "Posisi", "Status Kerja", "Sosmed Instansi",
    ]

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row_num, r in enumerate(records, 2):
        c = r.contact
        k = r.career
        ws.append([
            r.nama, r.nim, r.tahun_masuk,
            r.tgl_lulus.strftime("%Y-%m-%d") if r.tgl_lulus else "",
            r.fakultas, r.prodi,
            c.email if c else "", c.no_hp if c else "",
            c.linkedin if c else "", c.instagram if c else "",
            c.facebook if c else "", c.tiktok if c else "",
            k.tempat_kerja if k else "", k.alamat_kerja if k else "",
            k.posisi if k else "",
            k.status_kerja.value if k and k.status_kerja else "",
            k.sosmed_instansi if k else "",
        ])

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    log_action(
        db, current_user, "EXPORT_EXCEL", "/export/excel",
        detail=f"fakultas={fakultas} prodi={prodi} tahun={tahun} rows={len(records)}",
        ip_address=request.client.host if request.client else None,
    )

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=alumni_export.xlsx"},
    )


# ─── GET /admin/logs ──────────────────────────────────────────────────────────

@router.get("/admin/logs", response_model=PaginatedAuditLog)
def get_audit_logs(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    action: Optional[str] = Query(None),
    username: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    query = db.query(AuditLog).order_by(AuditLog.timestamp.desc())

    if action:
        query = query.filter(AuditLog.action.ilike(f"%{action}%"))
    if username:
        query = query.filter(AuditLog.username.ilike(f"%{username}%"))

    total = query.count()
    total_pages = max(1, math.ceil(total / limit))
    page = min(page, total_pages)
    records = query.offset((page - 1) * limit).limit(limit).all()

    return PaginatedAuditLog(
        data=records,
        total=total,
        page=page,
        total_pages=total_pages,
        limit=limit,
    )
